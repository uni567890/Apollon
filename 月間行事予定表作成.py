import streamlit as st
import csv
import openpyxl
from datetime import datetime
import io

def write_to_excel(csv_file_content, excel_file_content, year, month):
    """CSVデータをExcelの月間行事予定表に書き込む関数"""
    try:
        # Excelファイルをアップロードされた内容から読み込む
        wb = openpyxl.load_workbook(io.BytesIO(excel_file_content))
        sheet = wb.active
        weekdays_jp = ["月", "火", "水", "木", "金", "土", "日"]

        csv_file = io.StringIO(csv_file_content.decode('utf-8-sig'))
        reader = csv.DictReader(csv_file)

        for row in reader:
            start_date_str = row['開始日']
            start_time_str = row['開始時刻']
            end_time_str = row['終了時刻']
            event_name = row['件名']
            location = row['場所']

            try:  # 日付フォーマットエラーを捕捉
                start_date = datetime.strptime(start_date_str, '%Y/%m/%d').date()
            except ValueError:
                st.warning(f"日付形式エラー: {start_date_str}. この行はスキップされます。")
                continue #次の行へ

            if start_date.year != year or start_date.month != month:
                continue

            day = start_date.day
            weekday_num = start_date.weekday()
            weekday_jp = weekdays_jp[weekday_num]

            try: #時刻フォーマットエラーを捕捉
                start_time = datetime.strptime(start_time_str, '%H:%M').strftime('%H:%M')
                end_time = datetime.strptime(end_time_str, '%H:%M').strftime('%H:%M')
            except ValueError:
                st.warning(f"時刻形式エラー: {start_time_str}または{end_time_str}. 時間は空白で出力します。")
                st.warning(f"問題の開始時刻文字列: {start_time_str}") # 追加
                st.warning(f"問題の終了時刻文字列: {end_time_str}")   # 追加
                start_time = ""
                end_time = ""

            time_range = f"{start_time}\n-{end_time}"
            
            if location:
                location = location.split(',')[0].strip()

            if day <= 16:
                row_num = day + 8
                if sheet.cell(row=row_num, column=3).value is None:
                    sheet.cell(row=row_num, column=3).value = day
                sheet.cell(row=row_num, column=4).value = weekday_jp
                sheet.cell(row=row_num, column=5).value = time_range
                sheet.cell(row=row_num, column=6).value = event_name
                sheet.cell(row=row_num, column=7).value = location
            else:
                row_num = day - 16 + 8
                if sheet.cell(row=row_num, column=8).value is None:
                    sheet.cell(row=row_num, column=8).value = day
                sheet.cell(row=row_num, column=9).value = weekday_jp
                sheet.cell(row=row_num, column=10).value = time_range
                sheet.cell(row=row_num, column=11).value = event_name
                sheet.cell(row=row_num, column=12).value = location

        return wb

    except Exception as e:
        st.error(f"エラーが発生しました：{e}")
        return None

st.title("CSV to Excel 行事予定表")

uploaded_csv = st.file_uploader("CSVファイルをアップロードしてください", type="csv")
uploaded_excel = st.file_uploader("Excelテンプレートファイルをアップロードしてください", type="xlsx") # Excelファイルアップローダーを追加
year = st.number_input("年", min_value=1900, max_value=2100, value=datetime.now().year)
month = st.number_input("月", min_value=1, max_value=12, value=datetime.now().month)

if uploaded_csv is not None and uploaded_excel is not None: # 両方のファイルがアップロードされた場合のみ処理
    if st.button("Excelファイル作成"):
        try:
            wb = write_to_excel(uploaded_csv.getvalue(), uploaded_excel.getvalue(), year, month)
            if wb:
                virtual_workbook = io.BytesIO()
                wb.save(virtual_workbook)
                virtual_workbook.seek(0)
                st.download_button(
                    label="Excelファイルをダウンロード",
                    data=virtual_workbook,
                    file_name=f"{year}_{month}_gyoujiyotei.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
        except Exception as e:
            st.error(f"エラーが発生しました：{e}")
elif uploaded_csv is None:
    st.warning("CSVファイルをアップロードしてください。")
elif uploaded_excel is None:
    st.warning("Excelテンプレートファイルをアップロードしてください。")
